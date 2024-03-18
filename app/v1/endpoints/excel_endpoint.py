from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import os
import uuid
import shutil
from jinja2 import Template

router = APIRouter()
EXCEL_WORKSPACE = os.path.abspath("app\\v1\static\excel\\")
BASE_URL = "http://192.168.1.44:8000/"


class ExcelFilePath(BaseModel):
    excel_file_1_path: str
    excel_file_1_sheet_number: int = 1
    excel_file_2_path: str
    excel_file_2_sheet_number: int = 1

class RemoveExcelSession(BaseModel):
    session_id: str

class ExcelComparator:
    def __init__(self, file_paths: ExcelFilePath):
        self.document_1 = file_paths.excel_file_1_path
        self.document_1_sheet_number = file_paths.excel_file_1_sheet_number
        self.document_2 = file_paths.excel_file_2_path
        self.document_2_sheet_number = file_paths.excel_file_2_sheet_number

    def validate_documents(self):
        if not os.path.isfile(self.document_1):
            raise HTTPException(status_code=404, detail=f"{self.document_1} not found.")
        if not os.path.isfile(self.document_2):
            raise HTTPException(status_code=404, detail=f"{self.document_2} not found.")

        if not self.validate_xlsx_format():
            raise HTTPException(status_code=400, detail="Invalid document format")

        if not self.validate_excel_sheet_number():
            raise HTTPException(status_code=400, detail="Invalid document sheet number")

        if self.is_excel_sheet_blank():
            raise HTTPException(status_code=400, detail="Document is blank")

    def validate_xlsx_format(self) -> bool:
        if not self.document_1.endswith(".xlsx") or not self.document_2.endswith(".xlsx"):
            return False
        try:
            load_workbook(self.document_1)
            load_workbook(self.document_2)
            return True
        except Exception as error:
            return False

    def validate_excel_sheet_number(self) -> bool:
        workbook_1 = load_workbook(self.document_1)
        workbook_2 = load_workbook(self.document_2)

        if self.document_1_sheet_number < 1 or self.document_1_sheet_number > len(workbook_1.sheetnames):
            return False
        if self.document_2_sheet_number < 1 or self.document_2_sheet_number > len(workbook_2.sheetnames):
            return False
        return True

    def is_excel_sheet_blank(self) -> bool:
        workbook_1 = load_workbook(self.document_1)
        workbook_2 = load_workbook(self.document_2)

        doc_1_sheet = workbook_1.worksheets[self.document_1_sheet_number - 1]
        doc_2_sheet = workbook_2.worksheets[self.document_2_sheet_number - 1]

        for row in doc_1_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    return False

        for row in doc_2_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    return False

        return True


class Workspace:
    @staticmethod
    def create_session_workspace():
        session_id = str(uuid.uuid4())
        session_folder = os.path.join(EXCEL_WORKSPACE, session_id)
        os.makedirs(session_folder)
        return session_id

    @staticmethod
    def copy_documents_to_session_workspace(file1_path: str, file2_path: str, session_workspace_path: str):
        try:
            shutil.copy(file1_path, session_workspace_path)
            shutil.copy(file2_path, session_workspace_path)
            return True
        except Exception as error:
            return False


class HtmlGenerator:
    @staticmethod
    def generate_html_file(session_path, title, file1, file1_sheet_number,
                           data1, file2, file2_sheet_number,
                           data2, highlighted_rows):
        template_str = '''<!DOCTYPE html>
        <html>
            <head>
                <title>{{ title }}</title>
                <!-- Bootstrap CSS -->
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-EVSTQN3/azprG1Anm3QDgpJLIm9Nao0Yz1ztcQTwFspd3yD65VohhpuuCOmLASjC" crossorigin="anonymous">
                <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.7.1/jquery.min.js"></script>
                <style>
                .table-responsive {
                    overflow-x: unset;
                }
                .table-responsive table {
                    width: 100%;
                }
                .table-container {
                    display: flex;
                }
                .table-container .table-responsive {
                    flex: 0 0 auto;
                    margin-right: 10px;
                }
                .square-badge {
                    border-radius: 0;
                }
                .divScrollDiv {
                    display: inline-block;
                    width: 100%;
                    border: 1px solid black;
                    height: 94vh;
                    overflow: scroll;
                }
                .tableNoScroll {
                    overflow: hidden;
                }
            </style>
            <script>
                $(document).ready(function () {
                    var target_sec = $("#divFixed");
                    $("#divLista").scroll(function () {
                        target_sec.prop("scrollTop", this.scrollTop)
                        .prop("scrollLeft", this.scrollLeft);
                    });
                    var target_first = $("#divLista");
                    $("#divFixed").scroll(function () {
                        target_first.prop("scrollTop", this.scrollTop)
                        .prop("scrollLeft", this.scrollLeft);
                    });
                });
            </script>
        </head>
        <body>
        <div class="col-lg mx-auto p-1 py-md-1">
            <header class="d-flex align-items-center pb-1">
                <a href="#" class="d-flex align-items-center text-dark text-decoration-none">
                    <img src="/static/images/logo.jpeg" width="32" height="32" class="p-1">
                    <span class="fs-6">Document Comparison and Analysis (Demo)</span>
                </a>
            </header>
            <div class="container-fluid">
                <div class="row">
                    <div class="col divScrollDiv border" id="divFixed">
                        <div class="table-container">
                            <!-- First Table -->
                            <div class="table table-responsive mt-2">
                                <span class="badge bg-primary square-badge">Excel Document Path</span><span class="badge bg-success square-badge">{{ file1 }}</span><br>
                                <span class="badge bg-primary square-badge">Excel Sheet Number</span><span class="badge bg-success square-badge">{{ file1_sheet_number }}</span>
                                <table class="table-sm table-bordered mt-2">
                                    <thead class="table-dark">
                                        <tr>
                                            {% for i in data1[0].keys() %}
                                                <th>{{ i }}</th>
                                            {% endfor %}
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for row in data1 %}
                                            <tr {% if loop.index0 in highlighted_rows %}class="table-warning"{% endif %}>
                                                {% for value in row.values() %}
                                                    <td>{{ value }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                    <div class="col divScrollDiv border" id="divLista">
                        <div class="table-container">
                            <div class="table table-responsive mt-2">
                                <!-- Second Table -->
                                <span class="badge bg-primary square-badge">Excel Document Path</span><span class="badge bg-success square-badge">{{ file2 }}</span><br>
                                <span class="badge bg-primary square-badge">Excel Sheet Number</span><span class="badge bg-success square-badge">{{ file2_sheet_number }}</span>
                                <table class="table-sm table-bordered mt-2">
                                    <thead class="table-dark">
                                        <tr>
                                            {% for i in data2[0].keys() %}
                                                <th>{{ i }}</th>
                                            {% endfor %}
                                        </tr>
                                    </thead>
                                    <tbody>
                                        {% for row in data2 %}
                                            <tr {% if loop.index0 in highlighted_rows %}class="table-warning"{% endif %}>
                                                {% for value in row.values() %}
                                                    <td>{{ value }}</td>
                                                {% endfor %}
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <!-- Bootstrap JS (optional) -->
        <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.0.2/dist/js/bootstrap.bundle.min.js" integrity="sha384-MrcW6ZMFYlzcLA8Nl+NtUVF0sA7MsXsP1UyJoMp4YLEuNSfAP+JcXn/tWtIaxVXM" crossorigin="anonymous"></script>
    </body>
    </html>'''
        template = Template(template_str)
        rendered_html = template.render(
            title=title,
            file1=file1,
            file1_sheet_number=file1_sheet_number,
            data1=data1,
            file2=file2,
            file2_sheet_number=file2_sheet_number,
            data2=data2,
            highlighted_rows=highlighted_rows
        )
        with open(f"{session_path}/comparison_result.html", "w") as file:
            file.write(rendered_html)

@router.post("/remove_excel_session")
async def remove_excel_session(sessionid: RemoveExcelSession):
    excel_session_id = sessionid.session_id
    
    """check if session workspace is available"""
    SESSION_WORKSPACE = os.path.join(EXCEL_WORKSPACE, excel_session_id)
    if not Path(SESSION_WORKSPACE):
        raise HTTPException(status_code=404, detail=f"Session id {excel_session_id} not available")
    else:
        data = {"message": f"Session id {excel_session_id} removed successfully."}
        try:
            shutil.rmtree(SESSION_WORKSPACE)
            return JSONResponse(content=data, status_code=200)
        except Exception as error:
            raise HTTPException(status_code=500, detail=f"Session id {excel_session_id} not available")
            
@router.post("/generate_url_for_excel_doc")
async def generate_url(file_paths: ExcelFilePath):
    comparator = ExcelComparator(file_paths)

    """validate document"""
    comparator.validate_documents()

    """create session workspace"""
    session_id = Workspace.create_session_workspace()
    session_workspace = os.path.join(EXCEL_WORKSPACE, session_id)

    if not Workspace.copy_documents_to_session_workspace(file_paths.excel_file_1_path,
                                                         file_paths.excel_file_2_path,
                                                         session_workspace):
        raise HTTPException(status_code=500, detail="Error while copying the documents to session workspace")

    new_doc1_path = os.path.join(session_workspace, os.path.basename(file_paths.excel_file_1_path))
    new_doc2_path = os.path.join(session_workspace, os.path.basename(file_paths.excel_file_2_path))

    df1 = pd.read_excel(new_doc1_path, sheet_name=file_paths.excel_file_1_sheet_number - 1)
    df2 = pd.read_excel(new_doc2_path, sheet_name=file_paths.excel_file_2_sheet_number - 1)

    """ensure both the dataframes have the same columns"""
    columns = list(set(df1.columns) | set(df2.columns))
    df1 = df1.reindex(columns=columns)
    df2 = df2.reindex(columns=columns)

    """find the length of each dataframe"""
    len_df1 = len(df1)
    len_df2 = len(df2)

    """check which DataFrame has more rows """
    if len_df1 > len_df2:
        """append missing rows in df2 with None values"""
        missing_rows = len_df1 - len_df2
        df2 = pd.concat([df2, pd.DataFrame([[None]*len(columns)]*missing_rows, columns=columns)], ignore_index=True)
    elif len_df2 > len_df1:
        """Append missing rows in df1 with None values"""
        missing_rows = len_df2 - len_df1
        df1 = pd.concat([df1, pd.DataFrame([[None]*len(columns)]*missing_rows, columns=columns)], ignore_index=True)
    
    """heightlight the rows"""
    diff_df = df1.compare(df2)
    highlighted_rows = diff_df.index.tolist()

    """get the indices of all rows that are different or present only in one dataframe"""
    highlighted_rows = list(set(diff_df.index).union(set(df1.index).symmetric_difference(set(df2.index))))

    """convert dataframes to dictionary"""
    table1 = df1.to_dict(orient='records')
    table2 = df2.to_dict(orient='records')

    """update table1"""
    new_table_1 = [convert_keys_to_strings(item) for item in table1]
    new_table_1 = update_data_with_default(new_table_1, "N/A")

    """update table2"""
    new_table_2 = [convert_keys_to_strings(item) for item in table2]
    new_table_2 = update_data_with_default(new_table_2, "N/A")

    """generate html"""
    generate_html = HtmlGenerator()
    generate_html.generate_html_file(session_workspace, "Contentverse Excel Document Comparision", file_paths.excel_file_1_path, 
                       file_paths.excel_file_1_sheet_number, new_table_1,
                        file_paths.excel_file_2_path, file_paths.excel_file_2_sheet_number, new_table_2, 
                       highlighted_rows)
    """generate URL"""
    comparison_result_url = f"{BASE_URL}static/excel/{session_id}/comparison_result.html"
    
    return {"session_id": session_id, "comparison_result_url": comparison_result_url}

"""func: convert integer keys to string"""
def convert_keys_to_strings(dictionary):
    updated_dict = {}
    for key, value in dictionary.items():
        if isinstance(key, int):
            updated_dict["ID"] = value
        else:
            updated_dict[key] = value
    return updated_dict

"""func: update data with default value"""
def update_data_with_default(data_list, default_value):
    updated_data = []
    for item in data_list:
        updated_item = {key: value if value is not None else default_value for key, value in item.items()}
        updated_data.append(updated_item)
    return updated_data
